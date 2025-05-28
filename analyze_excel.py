import openpyxl
import os
from openpyxl.utils.cell import get_column_letter

# カレントディレクトリを表示
print(f"現在のディレクトリ: {os.getcwd()}")

try:
    # 計算式を取得するためのワークブック
    file_path = "各計算式 Ver 1.1.xlsx"
    print(f"ファイルを開こうとしています: {file_path}")
    
    wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
    
    # 値を取得するためのワークブック
    wb_values = openpyxl.load_workbook(file_path, data_only=True)
    
    # シート名を表示
    print(f"シート名: {wb_formulas.sheetnames}")
    
    # 分析結果を保存する辞書
    analysis_data = {}
    
    # 各シートの内容を分析
    for sheet_name in wb_formulas.sheetnames:
        print(f"\n==== シート: {sheet_name} ====")
        ws_formulas = wb_formulas[sheet_name]
        ws_values = wb_values[sheet_name]
        
        # 分析データ収集
        cell_data = {}
        
        # 値と計算式のあるセルの範囲を特定
        min_row = ws_formulas.min_row
        max_row = ws_formulas.max_row
        min_col = ws_formulas.min_column
        max_col = ws_formulas.max_column
        
        # ヘッダー行と列を取得する
        headers_row = {}
        headers_col = {}
        
        for row in range(min_row, min_row + 3):  # 最初の数行をヘッダーと仮定
            for col in range(min_col, max_col + 1):
                cell_value = ws_values.cell(row=row, column=col).value
                if cell_value:
                    col_letter = get_column_letter(col)
                    headers_row[col_letter] = cell_value
        
        for col in range(min_col, min_col + 3):  # 最初の数列をヘッダーと仮定
            col_letter = get_column_letter(col)
            for row in range(min_row, max_row + 1):
                cell_value = ws_values.cell(row=row, column=col).value
                if cell_value:
                    headers_col[f"{col_letter}{row}"] = cell_value
        
        # すべてのセルを調査
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell_formula = ws_formulas.cell(row=row, column=col)
                cell_value = ws_values.cell(row=row, column=col)
                
                col_letter = get_column_letter(col)
                cell_coord = f"{col_letter}{row}"
                
                # 値または計算式があるセルのみ保存
                if cell_formula.value or cell_value.value:
                    formula = cell_formula.value if cell_formula.data_type == 'f' else None
                    value = cell_value.value
                    
                    # 関連するヘッダーを探す
                    row_header = headers_row.get(col_letter, "")
                    col_header = ""
                    for i in range(1, 4):  # 左の数列をチェック
                        check_col = get_column_letter(col - i) if col - i >= min_col else None
                        if check_col:
                            check_coord = f"{check_col}{row}"
                            if check_coord in headers_col:
                                col_header = headers_col[check_coord]
                                break
                    
                    cell_data[cell_coord] = {
                        "値": value,
                        "計算式": formula,
                        "列ヘッダー": row_header,
                        "行ヘッダー": col_header
                    }
        
        analysis_data[sheet_name] = cell_data
    
    # マークダウン形式で結果を出力
    print("\n\n# Excelファイル分析結果\n")
    
    for sheet_name, cells in analysis_data.items():
        print(f"## シート: {sheet_name}\n")
        
        # 計算式のあるセルを先に出力
        print("### 計算式が設定されているセル\n")
        print("| セル | 値 | 計算式 | 推測される意味 |")
        print("|------|----|---------|--------------------|")
        
        for coord, data in cells.items():
            if data["計算式"]:
                formula = data["計算式"]
                value = data["値"]
                row_header = data["列ヘッダー"]
                col_header = data["行ヘッダー"]
                
                meaning = ""
                if row_header or col_header:
                    if row_header and col_header:
                        meaning = f"{col_header}の{row_header}"
                    elif row_header:
                        meaning = row_header
                    elif col_header:
                        meaning = f"{col_header}の値"
                
                print(f"| {coord} | {value} | {formula} | {meaning} |")
        
        # 値だけが入力されているセルを出力
        print("\n### 値のみが入力されているセル\n")
        print("| セル | 値 | 推測される意味 |")
        print("|------|----|-----------------|")
        
        for coord, data in cells.items():
            if not data["計算式"] and data["値"] is not None:
                value = data["値"]
                row_header = data["列ヘッダー"]
                col_header = data["行ヘッダー"]
                
                meaning = ""
                if row_header or col_header:
                    if row_header and col_header:
                        meaning = f"{col_header}の{row_header}"
                    elif row_header:
                        meaning = row_header
                    elif col_header:
                        meaning = f"{col_header}の値"
                
                print(f"| {coord} | {value} | {meaning} |")
    
except Exception as e:
    print(f"エラーが発生しました: {e}") 