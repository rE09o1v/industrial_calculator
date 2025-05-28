import openpyxl
import os

# カレントディレクトリを表示
print(f"現在のディレクトリ: {os.getcwd()}")

try:
    # ファイルを開く
    file_path = "各計算式 Ver 1.1.xlsx"
    print(f"ファイルを開こうとしています: {file_path}")
    
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    # シート名を表示
    print(f"シート名: {wb.sheetnames}")
    
    # 各シートの計算式を抽出
    for sheet_name in wb.sheetnames:
        print(f"\n==== シート: {sheet_name} ====")
        ws = wb[sheet_name]
        
        formula_count = 0
        # 全セルを走査して計算式を抽出
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # 計算式セル
                    formula_count += 1
                    print(f"{cell.coordinate}: {cell.value}")
        
        if formula_count == 0:
            print("このシートには計算式が見つかりませんでした")
    
except Exception as e:
    print(f"エラーが発生しました: {e}") 